from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()
    #print(len(sheets))
    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表中所有数据
        sheetInfo = []
        for lineNum in range(1,sheet.max_row+1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
        dic[sheetName] = sheetInfo
    return dic

path = ''# .xlsx
dic = readXlsxFile(path)
print(dic)
print(len(dic))