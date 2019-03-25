# xlsx xls
# openpyxl模块->xlsx
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):

    file = load_workbook(filename=path)  #打开文件
    sheets = file.get_sheet_names()  #所有表格的名称
    sheet = file.get_sheet_by_name(sheets[0]) #拿出一个表格 sheets[0]
    #print(sheet.max_row)      #最大行数
    #print(sheet.max_column)   #最大列数
    #print(sheet.title)        #表名



    for lineNum in range(1, sheet.max_row+1):
        #print(lineNum)
        lineList = []
        #print(sheet.max_row, sheet.max_column)
        for columnNum in range(1, sheet.max_column+1):
            #拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            #if value != None:
            lineList.append(value)
        print(lineList)  #打印一行数据
#不能读取xls文件
path = ''# .xlsx
readXlsxFile(path)